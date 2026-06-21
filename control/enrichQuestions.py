import os
import sys
import time
import json

# Setup paths
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL_PATH = os.path.abspath(os.path.join(BASE_DIR, 'symlinks', 'ObjectiveQuestions.xlsx'))
TAG_RULES_PATH = os.path.join(BASE_DIR, 'tag_rules.json')

# Dependency checks
try:
    import openpyxl
except ImportError:
    print("Error: Missing 'openpyxl' package. Run: pip install openpyxl")
    sys.exit(1)

try:
    import google.generativeai as genai
except ImportError:
    print("Error: Missing 'google-generativeai' package. Run: pip install google-generativeai")
    sys.exit(1)

def load_tag_rules():
    if not os.path.exists(TAG_RULES_PATH):
        print(f"Warning: tag_rules.json not found at {TAG_RULES_PATH}")
        return {}
    try:
        with open(TAG_RULES_PATH, 'r', encoding='utf-8') as f:
            return json.load(f)
    except Exception as e:
        print(f"Error loading tag_rules.json: {e}")
        return {}

def main():
    print("[QuizMe Question Enricher - Gemini AI Studio Free Tier]")
    print("=========================================================")

    # Get API Key (supports .env files via python-dotenv)
    try:
        from dotenv import load_dotenv
        load_dotenv(os.path.join(BASE_DIR, '.env'))
    except ImportError:
        pass

    api_key = os.environ.get("GEMINI_API_KEY")
    if not api_key:
        print("GEMINI_API_KEY environment variable not set.")
        api_key = input("Please enter your Gemini API Key from Google AI Studio: ").strip()
        if not api_key:
            print("Error: Gemini API Key is required to run this script.")
            sys.exit(1)
            
    genai.configure(api_key=api_key)
    
    # Configure model to output JSON
    generation_config = {
        "response_mime_type": "application/json"
    }
    
    try:
        model = genai.GenerativeModel("gemini-1.5-flash", generation_config=generation_config)
    except Exception as e:
        print(f"Error initializing Gemini model: {e}")
        sys.exit(1)

    # Load tag rules for subtopic mapping
    tag_rules = load_tag_rules()
    allowed_tags = list(tag_rules.keys())
    
    # Load Excel Workbook
    if not os.path.exists(EXCEL_PATH):
        print(f"Error: Excel file not found at {EXCEL_PATH}")
        sys.exit(1)
        
    print(f"Opening workbook: {EXCEL_PATH}")
    try:
        wb = openpyxl.load_workbook(EXCEL_PATH)
    except Exception as e:
        print(f"Error opening workbook: {e}")
        sys.exit(1)
        
    expected_cols = ['QID', 'Question', 'Answer', 'Choice1', 'Choice2', 'Choice3', 'Information', 'Tags']
    
    enriched_count = 0
    
    for sheet in wb.worksheets:
        sheet_name = sheet.title.strip()
        print(f"\nScanning sheet: '{sheet_name}'...")
        
        # Determine column map
        headers = [str(sheet.cell(1, c).value).strip() if sheet.cell(1, c).value else ''
                   for c in range(1, sheet.max_column + 1)]
        col_map = {h: i+1 for i, h in enumerate(headers)}
        
        # Ensure expected columns exist
        for col_name in expected_cols:
            if col_name not in col_map:
                new_col = sheet.max_column + 1
                sheet.cell(1, new_col).value = col_name
                col_map[col_name] = new_col
                print(f"  Added missing column: {col_name}")
                
        # Find rows to enrich
        q_col = col_map['Question']
        a_col = col_map['Answer']
        c1_col = col_map['Choice1']
        c2_col = col_map['Choice2']
        c3_col = col_map['Choice3']
        info_col = col_map['Information']
        tags_col = col_map['Tags']
        
        for row in range(2, sheet.max_row + 1):
            q_val = sheet.cell(row, q_col).value
            a_val = sheet.cell(row, a_col).value
            
            question_text = str(q_val).strip() if q_val else ''
            answer_text = str(a_val).strip() if a_val else ''
            
            if question_text and not answer_text:
                print(f"  Row {row}: Found question needing options...")
                print(f"    Q: \"{question_text[:80]}...\"")
                
                # Construct AI prompt
                prompt = f"""
                You are an expert tutor for the BPSC (Bihar Public Service Commission) and UPSC Civil Services prelims examinations.
                You are given the following question from the syllabus subject '{sheet_name}':
                
                Question: "{question_text}"
                
                Please enrich this question by generating:
                1. The correct answer.
                2. Three realistic, high-quality, and plausible distractors (incorrect options) for Choice1, Choice2, and Choice3. They should be challenging, typical of civil service exam standards, and not obviously fake.
                3. High-quality fact-checked explanation (Information) of the answer, suitable for a student studying for this exam (1-3 sentences).
                4. Relevant sub-topic tags from this allowed list: {allowed_tags}. Return them as a comma-separated string.
                
                Return a single JSON object matching this structure EXACTLY:
                {{
                  "answer": "Correct option text",
                  "choice1": "First incorrect option text",
                  "choice2": "Second incorrect option text",
                  "choice3": "Third incorrect option text",
                  "information": "Explanation text",
                  "tags": "tag1, tag2"
                }}
                """
                
                try:
                    response = model.generate_content(prompt)
                    ai_data = json.loads(response.text.strip())
                    
                    # Write back to Excel
                    sheet.cell(row, a_col).value = ai_data.get("answer", "").strip()
                    sheet.cell(row, c1_col).value = ai_data.get("choice1", "").strip()
                    sheet.cell(row, c2_col).value = ai_data.get("choice2", "").strip()
                    sheet.cell(row, c3_col).value = ai_data.get("choice3", "").strip()
                    
                    # Update explanation if empty
                    if not sheet.cell(row, info_col).value:
                        sheet.cell(row, info_col).value = ai_data.get("information", "").strip()
                        
                    # Update tags if empty
                    if not sheet.cell(row, tags_col).value:
                        sheet.cell(row, tags_col).value = ai_data.get("tags", "").strip()
                        
                    print(f"    [OK] AI response mapped: Answer=\"{ai_data.get('answer')}\"")
                    enriched_count += 1
                    
                    # Save after each enrichment to prevent data loss
                    wb.save(EXCEL_PATH)
                    
                    # Rate limiting: 4s sleep to remain under 15 RPM
                    time.sleep(4)
                    
                except Exception as e:
                    print(f"    [ERROR] Error generating or writing options: {e}")
                    
    print(f"\nFinished! Enriched {enriched_count} question(s) successfully.")

if __name__ == "__main__":
    main()
