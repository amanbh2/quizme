# ═══════════════════════════════════════════════════════════════
#  excel_to_csv.py  —  QuizMe v5.0 (Excel -> CSV Converter)
#  Converts .xlsx sheets in control/symlinks/ to individual CSVs
# ═══════════════════════════════════════════════════════════════

import os
import csv
import sys

try:
    import openpyxl
except ImportError:
    print("Missing dependency. Run: pip install openpyxl")
    sys.exit(1)

# ── Paths ──────────────────────────────────────────────────────
BASE_DIR     = os.path.dirname(os.path.abspath(__file__))
SYMLINKS_DIR = os.path.join(BASE_DIR, 'symlinks')

def convert_excel_to_csv():
    if not os.path.exists(SYMLINKS_DIR):
        print(f"Directory not found: {SYMLINKS_DIR}")
        return

    # Find all .xlsx files in control/symlinks
    xlsx_files = [f for f in os.listdir(SYMLINKS_DIR) if f.endswith('.xlsx')]
    if not xlsx_files:
        print("No Excel files found in control/symlinks/")
        return

    for f in xlsx_files:
        xlsx_path = os.path.join(SYMLINKS_DIR, f)
        print(f"\nProcessing workbook: {f}")
        try:
            wb = openpyxl.load_workbook(xlsx_path, data_only=True)
        except Exception as e:
            print(f"  ✗ Failed to load workbook: {e}")
            continue

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # Normalize sheet name to create CSV filename (e.g. 'Bihar Economy' -> 'BiharEconomy.csv')
            csv_filename = sheet_name.replace(' ', '') + '.csv'
            csv_path = os.path.join(SYMLINKS_DIR, csv_filename)

            rows_written = 0
            with open(csv_path, mode='w', encoding='utf-8-sig', newline='') as csv_file:
                writer = csv.writer(csv_file)
                for r in range(1, ws.max_row + 1):
                    row_vals = []
                    for c in range(1, ws.max_column + 1):
                        v = ws.cell(r, c).value
                        if v is None:
                            row_vals.append('')
                        else:
                            # Fix whole numbers stored as floats (e.g., 943.0 -> "943")
                            if isinstance(v, float) and v == int(v):
                                v = int(v)
                            row_vals.append(str(v).strip())
                    
                    # Skip completely empty rows
                    if any(val != '' for val in row_vals):
                        writer.writerow(row_vals)
                        rows_written += 1
            
            # If nothing was written (empty sheet or only header), clean up the file
            if rows_written <= 1:
                if os.path.exists(csv_path):
                    os.remove(csv_path)
            else:
                print(f"  ✓ Sheet '{sheet_name}' -> {csv_filename} ({rows_written} rows)")

if __name__ == "__main__":
    convert_excel_to_csv()
