# ═══════════════════════════════════════════════════════════════
#  excel_to_csv.py  —  QuizMe v5.0 (Excel -> CSV Converter)
#  Interactively converts .xlsx sheets in symlinks/ to individual CSVs
# ═══════════════════════════════════════════════════════════════

import os
import csv
import sys

try:
    import openpyxl
except ImportError:
    print("Missing dependency. Run: pip install openpyxl")
    sys.exit(1)

# ── Paths ──────────────────────────────────────────────────────
BASE_DIR     = os.path.dirname(os.path.abspath(__file__))
# symlinks directory is located in the parent directory (control/symlinks)
SYMLINKS_DIR = os.path.abspath(os.path.join(BASE_DIR, '..', 'symlinks'))

def select_from_list(prompt, items, allow_cancel=True):
    while True:
        print(f"\n{prompt}")
        for idx, item in enumerate(items, 1):
            print(f"  [{idx}] {item}")
        if allow_cancel:
            print("  [q] Quit")
        
        choice = input(f"Select an option (1-{len(items)} or 'q'): ").strip()
        if choice.lower() == 'q' and allow_cancel:
            print("Exiting.")
            sys.exit(0)
        try:
            idx = int(choice)
            if 1 <= idx <= len(items):
                return items[idx - 1]
            else:
                print(f"✗ Invalid choice. Please enter a number between 1 and {len(items)}.")
        except ValueError:
            print("✗ Invalid input. Please enter a number or 'q'.")

def convert_excel_to_csv():
    if not os.path.exists(SYMLINKS_DIR):
        print(f"Directory not found: {SYMLINKS_DIR}")
        return

    # Find all .xlsx files in control/symlinks
    xlsx_files = [f for f in os.listdir(SYMLINKS_DIR) if f.endswith('.xlsx')]
    if not xlsx_files:
        print(f"No Excel files found in {SYMLINKS_DIR}")
        return

    # 1. Select workbook
    selected_workbook = select_from_list("Available Workbooks:", xlsx_files)
    xlsx_path = os.path.join(SYMLINKS_DIR, selected_workbook)

    print(f"\nLoading workbook: {selected_workbook}...")
    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    except Exception as e:
        print(f"  ✗ Failed to load workbook: {e}")
        return

    # 2. Select sheet
    sheet_names = wb.sheetnames
    if not sheet_names:
        print("✗ No sheets found in workbook.")
        return

    selected_sheet = select_from_list(f"Available Sheets in {selected_workbook}:", sheet_names)
    ws = wb[selected_sheet]

    # Normalize sheet name to create CSV filename
    csv_filename = selected_sheet.replace(' ', '') + '.csv'
    csv_path = os.path.join(BASE_DIR, csv_filename)

    print(f"\nExtracting sheet '{selected_sheet}' to {csv_filename}...")
    rows_written = 0
    with open(csv_path, mode='w', encoding='utf-8-sig', newline='') as csv_file:
        writer = csv.writer(csv_file)
        for r in range(1, ws.max_row + 1):
            row_vals = []
            for c in range(1, ws.max_column + 1):
                v = ws.cell(r, c).value
                if v is None:
                    row_vals.append('')
                else:
                    # Fix whole numbers stored as floats (e.g., 943.0 -> "943")
                    if isinstance(v, float) and v == int(v):
                        v = int(v)
                    row_vals.append(str(v).strip())
            
            # Skip completely empty rows
            if any(val != '' for val in row_vals):
                writer.writerow(row_vals)
                rows_written += 1
    
    # If nothing was written (empty sheet or only header), clean up the file
    if rows_written <= 1:
        if os.path.exists(csv_path):
            os.remove(csv_path)
        print("✗ Sheet was empty or contained only a header. No CSV written.")
    else:
        print(f"✓ Success! Generated: {csv_path} ({rows_written} rows)")

if __name__ == "__main__":
    convert_excel_to_csv()
