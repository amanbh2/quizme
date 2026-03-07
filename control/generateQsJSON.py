# ═══════════════════════════════════════════════════════════════
#  generateQsJSON.py  —  QuizMe v5.0
#  Merges: generateQs.py + createManifestFile.py
#  Modes: convert | autotag | stats | qid-report | renumber
# ═══════════════════════════════════════════════════════════════

import os, json, re, sys
from collections import Counter

try:
    import openpyxl
except ImportError:
    print("Missing dependency. Run: pip install openpyxl")
    sys.exit(1)

# ── Paths ──────────────────────────────────────────────────────
BASE_DIR      = os.path.dirname(os.path.abspath(__file__))
EXCEL_PATH    = r"C:\Users\amanb\OneDrive\Documents\ObjectiveQuestions.xlsx"
DATA_DIR      = os.path.join(BASE_DIR, '..', 'data')
CONTROL_DIR   = BASE_DIR
MANIFEST_PATH = os.path.join(CONTROL_DIR, 'manifest.json')
COUNTER_PATH  = os.path.join(CONTROL_DIR, 'qid_counter.json')
TAG_RULES_PATH= os.path.join(CONTROL_DIR, 'tag_rules.json')

# ── QID helpers ────────────────────────────────────────────────
QID_PATTERN = re.compile(r'^Q\d{5}$')

def is_valid_qid(val):
    return bool(val and QID_PATTERN.match(str(val).strip()))

def format_qid(n):
    return f"Q{n:05d}"

def load_counter(wb_qids):
    """Load counter — always rebuild from Excel max to be safe."""
    existing_nums = [int(q[1:]) for q in wb_qids if is_valid_qid(q)]
    max_from_excel = max(existing_nums, default=0)
    if os.path.exists(COUNTER_PATH):
        try:
            saved = json.load(open(COUNTER_PATH))
            max_from_file = saved.get('last', 0)
        except Exception:
            max_from_file = 0
    else:
        max_from_file = 0
    return max(max_from_excel, max_from_file)

def save_counter(n):
    json.dump({'last': n}, open(COUNTER_PATH, 'w'))

# ── Tag helpers ────────────────────────────────────────────────
def load_tag_rules():
    if not os.path.exists(TAG_RULES_PATH):
        print(f"  ⚠  tag_rules.json not found at {TAG_RULES_PATH}")
        return {}
    return json.load(open(TAG_RULES_PATH, encoding='utf-8'))

def auto_tag_question(text, info, tag_rules):
    """Return comma-separated tags for a question based on rules."""
    combined = (str(text) + ' ' + str(info)).lower()
    matched  = []
    for tag, keywords in tag_rules.items():
        for kw in keywords:
            if kw.lower() in combined:
                matched.append(tag)
                break
    return ','.join(matched)

# ── Excel helpers ──────────────────────────────────────────────
EXPECTED_COLS = ['QID', 'Question', 'Answer', 'Choice1', 'Choice2', 'Choice3', 'Information', 'Tags']

def get_col_map(sheet):
    """Return dict of col_name → col_index (1-based) from header row."""
    headers = [str(sheet.cell(1, c).value).strip() if sheet.cell(1, c).value else ''
               for c in range(1, sheet.max_column + 1)]
    return {h: i+1 for i, h in enumerate(headers)}

def ensure_columns(sheet):
    """Add any missing columns to header row."""
    col_map = get_col_map(sheet)
    for col_name in EXPECTED_COLS:
        if col_name not in col_map:
            new_col = sheet.max_column + 1
            sheet.cell(1, new_col).value = col_name
            col_map[col_name] = new_col
            print(f"  ℹ  Added missing column: {col_name}")
    return get_col_map(sheet)

def is_blank_row(sheet, row, col_map):
    """True if Question column is blank — skip this row."""
    q_col = col_map.get('Question')
    if not q_col:
        return True
    val = sheet.cell(row, q_col).value
    return not val or str(val).strip() == ''

# ── MANIFEST ───────────────────────────────────────────────────
def create_manifest():
    files = [f for f in os.listdir(DATA_DIR)
             if f.endswith('.json') and os.path.isfile(os.path.join(DATA_DIR, f))]
    files.sort()
    if os.path.exists(MANIFEST_PATH):
        os.remove(MANIFEST_PATH)
    json.dump({"files": files}, open(MANIFEST_PATH, 'w', encoding='utf-8'), indent=2)
    print(f"  ✓  Manifest updated — {len(files)} file(s)")

# ═══════════════════════════════════════════════════════════════
#  MODE: CONVERT
# ═══════════════════════════════════════════════════════════════
def mode_convert():
    print("\n── CONVERT ────────────────────────────────────────────")
    if not os.path.exists(EXCEL_PATH):
        print(f"  ✗  Excel not found: {EXCEL_PATH}"); return

    wb = openpyxl.load_workbook(EXCEL_PATH)
    os.makedirs(DATA_DIR, exist_ok=True)

    # Collect all existing QIDs across workbook first
    all_existing_qids = set()
    for ws in wb.worksheets:
        cm = get_col_map(ws)
        if 'QID' not in cm: continue
        for row in range(2, ws.max_row + 1):
            val = ws.cell(row, cm['QID']).value
            if is_valid_qid(val):
                all_existing_qids.add(str(val).strip())

    counter = load_counter(all_existing_qids)
    sheet_data = {}   # sheet_name → list of question dicts
    assigned   = 0
    duplicates = 0
    seen_qids  = set()

    for ws in wb.worksheets:
        sheet_name = ws.title.strip()
        col_map    = ensure_columns(ws)
        questions  = []

        for row in range(2, ws.max_row + 1):
            if is_blank_row(ws, row, col_map):
                continue

            # Read QID
            raw_qid = ws.cell(row, col_map['QID']).value
            raw_qid = str(raw_qid).strip() if raw_qid else ''

            # Validate / normalise
            if not is_valid_qid(raw_qid):
                raw_qid = ''  # treat as blank

            # Duplicate detection
            if raw_qid and raw_qid in seen_qids:
                print(f"  ⚠  Duplicate QID {raw_qid} in sheet '{sheet_name}' row {row} — reassigning")
                raw_qid = ''
                duplicates += 1

            # Assign new QID if blank
            if not raw_qid:
                counter += 1
                raw_qid  = format_qid(counter)
                ws.cell(row, col_map['QID']).value = raw_qid
                assigned += 1

            seen_qids.add(raw_qid)

            def cell(col):
                v = ws.cell(row, col_map.get(col, 0)).value if col_map.get(col) else None
                return str(v).strip() if v is not None else ''

            question = cell('Question')
            answer   = cell('Answer')
            choices  = [answer] + [c for c in [cell('Choice1'), cell('Choice2'),
                                     cell('Choice3')] if c]
            info     = cell('Information')
            tags     = cell('Tags')

            if not question or not answer:
                continue

            questions.append({
                "qid":         raw_qid,
                "question":    question,
                "answer":      answer,
                "choices":     choices,
                "information": info,
                "tags":        tags
            })

        sheet_data[sheet_name] = questions

    # Save counter
    save_counter(counter)

    # Write JSON files
    all_questions = []
    for sheet_name, questions in sheet_data.items():
        if not questions: continue
        fname = sheet_name.replace(' ', '') + '.json'
        fpath = os.path.join(DATA_DIR, fname)
        json.dump(questions, open(fpath, 'w', encoding='utf-8'), ensure_ascii=False, indent=2)
        all_questions.extend(questions)
        print(f"  ✓  {fname} — {len(questions)} questions")

    # Write all.json
    if all_questions:
        json.dump(all_questions,
                  open(os.path.join(DATA_DIR, 'all.json'), 'w', encoding='utf-8'),
                  ensure_ascii=False, indent=2)
        print(f"  ✓  all.json — {len(all_questions)} questions total")

    # Save workbook (QIDs written back)
    wb.save(EXCEL_PATH)
    print(f"\n  ✓  {assigned} new QID(s) assigned")
    if duplicates:
        print(f"  ⚠  {duplicates} duplicate QID(s) detected and reassigned")

    create_manifest()
    print("\n  ✓  Convert complete\n")

# ═══════════════════════════════════════════════════════════════
#  MODE: AUTOTAG
# ═══════════════════════════════════════════════════════════════
def mode_autotag():
    print("\n── AUTOTAG ─────────────────────────────────────────────")
    if not os.path.exists(EXCEL_PATH):
        print(f"  ✗  Excel not found: {EXCEL_PATH}"); return

    tag_rules = load_tag_rules()
    if not tag_rules:
        print("  ✗  No tag rules loaded. Check tag_rules.json"); return

    wb      = openpyxl.load_workbook(EXCEL_PATH)
    tagged  = 0
    skipped = 0

    for ws in wb.worksheets:
        col_map = ensure_columns(ws)
        if 'Tags' not in col_map or 'Question' not in col_map:
            continue

        for row in range(2, ws.max_row + 1):
            if is_blank_row(ws, row, col_map):
                continue

            existing_tag = ws.cell(row, col_map['Tags']).value
            # Never overwrite manually set tags
            if existing_tag and str(existing_tag).strip():
                skipped += 1
                continue

            question = ws.cell(row, col_map['Question']).value or ''
            info     = ws.cell(row, col_map.get('Information', 0)).value or '' \
                       if col_map.get('Information') else ''

            new_tags = auto_tag_question(question, info, tag_rules)
            if new_tags:
                ws.cell(row, col_map['Tags']).value = new_tags
                tagged += 1

    wb.save(EXCEL_PATH)
    print(f"  ✓  {tagged} question(s) tagged")
    print(f"  ℹ  {skipped} question(s) skipped (already have tags)")
    print("\n  ✓  Autotag complete — review tags in Excel before running convert\n")

# ═══════════════════════════════════════════════════════════════
#  MODE: STATS
# ═══════════════════════════════════════════════════════════════
def mode_stats():
    print("\n── STATS ───────────────────────────────────────────────")
    if not os.path.exists(DATA_DIR):
        print("  ✗  data/ folder not found. Run convert first."); return

    total = 0
    for f in sorted(os.listdir(DATA_DIR)):
        if not f.endswith('.json') or f == 'all.json': continue
        path = os.path.join(DATA_DIR, f)
        try:
            data = json.load(open(path, encoding='utf-8'))
            print(f"  {f.replace('.json',''):<30} {len(data):>5} questions")
            total += len(data)
        except Exception as e:
            print(f"  ✗  {f}: {e}")
    print(f"\n  {'TOTAL':<30} {total:>5} questions\n")

# ═══════════════════════════════════════════════════════════════
#  MODE: QID-REPORT
# ═══════════════════════════════════════════════════════════════
def mode_qid_report():
    print("\n── QID REPORT ──────────────────────────────────────────")
    if not os.path.exists(EXCEL_PATH):
        print(f"  ✗  Excel not found: {EXCEL_PATH}"); return

    wb   = openpyxl.load_workbook(EXCEL_PATH)
    qids = []

    for ws in wb.worksheets:
        col_map = get_col_map(ws)
        if 'QID' not in col_map: continue
        for row in range(2, ws.max_row + 1):
            val = ws.cell(row, col_map['QID']).value
            if is_valid_qid(val):
                qids.append(str(val).strip())

    if not qids:
        print("  ℹ  No valid QIDs found. Run convert first.\n"); return

    nums    = sorted([int(q[1:]) for q in qids])
    max_num = nums[-1]
    active  = len(nums)
    full    = set(range(1, max_num + 1))
    gaps    = sorted(full - set(nums))

    print(f"\n  Max QID reached : Q{max_num:05d}")
    print(f"  Total assigned  : {max_num}")
    print(f"  Currently active: {active}")
    print(f"  Gaps (deleted)  : {len(gaps)}")

    # Visual bar
    bar_len    = 40
    filled     = round((active / max_num) * bar_len)
    gap_filled = bar_len - filled
    print(f"\n  Q00001 {'█' * filled}{'░' * gap_filled} Q{max_num:05d}")
    print(f"         ■ {active} Active   □ {len(gaps)} Gaps\n")

    if gaps:
        gap_strs = [f"Q{g:05d}" for g in gaps]
        # Print in rows of 8
        print("  Gap QIDs:")
        for i in range(0, len(gap_strs), 8):
            print("    " + "  ".join(gap_strs[i:i+8]))
    print()

# ═══════════════════════════════════════════════════════════════
#  MODE: RENUMBER
# ═══════════════════════════════════════════════════════════════
def mode_renumber():
    print("\n── RENUMBER ────────────────────────────────────────────")
    print("""
  ⚠  FULL RENUMBER — this will change ALL QIDs in your Excel file.
  ─────────────────────────────────────────────────────────────────
  All existing QIDs will be reassigned sequentially from Q00001.

  YOU MUST manually reset statistics after this operation:
    → Open QuizMe → Stats tab → Reset All Statistics
    OR the app will detect orphaned stats and prompt you.

  Your Gist backup will also be outdated after renumber.
  ─────────────────────────────────────────────────────────────────
    """)
    confirm = input("  Type YES to proceed (anything else cancels): ").strip()
    if confirm != 'YES':
        print("  Cancelled.\n"); return

    if not os.path.exists(EXCEL_PATH):
        print(f"  ✗  Excel not found: {EXCEL_PATH}"); return

    wb      = openpyxl.load_workbook(EXCEL_PATH)
    counter = 0

    for ws in wb.worksheets:
        col_map = ensure_columns(ws)
        if 'QID' not in col_map: continue

        for row in range(2, ws.max_row + 1):
            if is_blank_row(ws, row, col_map): continue
            counter += 1
            ws.cell(row, col_map['QID']).value = format_qid(counter)

    save_counter(counter)
    wb.save(EXCEL_PATH)

    print(f"\n  ✓  Renumber complete — {counter} questions renumbered Q00001–Q{counter:05d}")
    print("""
  ══════════════════════════════════════════════════════
  ⚠  ACTION REQUIRED IN QUIZME:
     Your statistics are now outdated.

     Open QuizMe → Stats tab → Reset All Statistics
     The app will detect and prompt you automatically.
  ══════════════════════════════════════════════════════
    """)
    # Re-run convert automatically
    run_convert = input("  Run convert now to update JSONs? (y/n): ").strip().lower()
    if run_convert == 'y':
        mode_convert()

# ═══════════════════════════════════════════════════════════════
#  MAIN MENU
# ═══════════════════════════════════════════════════════════════
def main():
    print("""
╔═══════════════════════════════════════╗
║       generateQsJSON.py  v5.0         ║
║       QuizMe Question Manager         ║
╚═══════════════════════════════════════╝

  What would you like to do?

  convert     →  Assign QIDs + export JSONs + update manifest
  autotag     →  Auto-tag untagged questions using tag_rules.json
  stats       →  Show question counts per sheet
  qid-report  →  Show QID health (max, active, gaps)
  renumber    →  Full renumber all QIDs from Q00001 (needs stats reset)
    """)

    choice = input("  Enter mode: ").strip().lower()
    modes  = {
        'convert':    mode_convert,
        'autotag':    mode_autotag,
        'stats':      mode_stats,
        'qid-report': mode_qid_report,
        'renumber':   mode_renumber,
    }
    if choice in modes:
        modes[choice]()
    else:
        print(f"\n  ✗  Unknown mode: '{choice}'. Choose from: {', '.join(modes.keys())}\n")

if __name__ == '__main__':
    main()