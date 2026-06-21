#!/usr/bin/env python3
"""
generateTimelineJSON.py
Reads BCE and CE sheets from History_Timelines_2026.xlsx
and outputs a unified data/timeline.json for the QuizMe app.

Usage:
    python control/generateTimelineJSON.py
"""

import openpyxl
import json
import os
import sys

EXCEL_PATH = os.path.join(os.path.dirname(__file__), 'symlinks', 'History_Timelines_2026.xlsx')
OUTPUT_PATH = os.path.join(os.path.dirname(__file__), '..', 'data', 'timeline', 'timeline.json')

def read_bce_sheet(wb):
    """Read BCE sheet: columns BCE-Start, BCE-End, Event, Age, Indian Rulers.
    We keep: yearStart (negative), yearEnd (negative or null), event.
    """
    ws = wb['BCE']
    events = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        # Columns: BCE-Start, BCE-End, Event, Age, Indian Rulers
        bce_start = row[0]   # Can be None
        bce_end = row[1]     # Usually present
        event_text = row[2]

        if not event_text or not str(event_text).strip():
            continue

        # Determine the primary year (use bce_end as the "point in time")
        # bce_end is the more specific date; bce_start marks range start if present
        year = None
        year_end = None

        if bce_end is not None:
            try:
                year = -abs(int(bce_end))  # Negative for BCE
            except (ValueError, TypeError):
                continue

        if bce_start is not None:
            try:
                year_end = -abs(int(bce_start))  # Earlier date (more negative)
            except (ValueError, TypeError):
                year_end = None

        if year is None:
            continue

        # Clean up event text: replace literal \n with space
        event_text = str(event_text).replace('\\n', ' ').replace('\n', ' ').strip()
        # Collapse multiple spaces
        event_text = ' '.join(event_text.split())

        events.append({
            'year': year,
            'yearEnd': year_end,
            'event': event_text
        })

    # Sort by year (most ancient first — most negative)
    events.sort(key=lambda e: e['year'])
    return events


def read_ce_sheet(wb):
    """Read CE sheet: columns Day, Month, Year, Event, Age.
    We keep: year, day (nullable), month (nullable), event.
    """
    ws = wb['CE']
    events = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        # Columns: Day, Month, Year, Event, Age
        day = row[0]
        month = row[1]
        year = row[2]
        event_text = row[3]

        if not event_text or not str(event_text).strip():
            continue
        if year is None:
            continue

        try:
            year = int(year)
        except (ValueError, TypeError):
            continue

        # Normalise day
        if day is not None:
            try:
                day = int(day)
            except (ValueError, TypeError):
                day = None

        # Normalise month (should be string like "January")
        if month is not None:
            month = str(month).strip()
            if not month or month.lower() == 'none':
                month = None

        # Clean up event text
        event_text = str(event_text).replace('\\n', ' ').replace('\n', ' ').strip()
        event_text = ' '.join(event_text.split())

        events.append({
            'year': year,
            'day': day,
            'month': month,
            'event': event_text
        })

    # Sort by year, then month (rough), then day
    month_order = {
        'January': 1, 'February': 2, 'March': 3, 'April': 4,
        'May': 5, 'June': 6, 'July': 7, 'August': 8,
        'September': 9, 'October': 10, 'November': 11, 'December': 12
    }

    def sort_key(e):
        m = month_order.get(e.get('month'), 0) if e.get('month') else 0
        d = e.get('day') or 0
        return (e['year'], m, d)

    events.sort(key=sort_key)
    return events


def assign_ids(bce_events, ce_events):
    """Assign sequential T00001-style IDs."""
    counter = 1
    for e in bce_events:
        e['id'] = f'T{counter:05d}'
        counter += 1
    for e in ce_events:
        e['id'] = f'T{counter:05d}'
        counter += 1
    return counter - 1


def main():
    if not os.path.exists(EXCEL_PATH):
        print(f'ERROR: Excel file not found at {EXCEL_PATH}')
        sys.exit(1)

    print(f'Reading: {EXCEL_PATH}')
    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)

    print('Processing BCE sheet...')
    bce_events = read_bce_sheet(wb)
    print(f'  -> {len(bce_events)} BCE events')

    print('Processing CE sheet...')
    ce_events = read_ce_sheet(wb)
    print(f'  -> {len(ce_events)} CE events')

    total = assign_ids(bce_events, ce_events)
    print(f'  -> {total} total events (T00001–T{total:05d})')

    wb.close()

    # Build output
    output = {
        'bce': bce_events,
        'ce': ce_events
    }

    # Ensure output directory exists
    os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)

    with open(OUTPUT_PATH, 'w', encoding='utf-8') as f:
        json.dump(output, f, ensure_ascii=False, indent=2)

    file_size = os.path.getsize(OUTPUT_PATH)
    print(f'\nGenerated: {OUTPUT_PATH}')
    print(f'File size: {file_size / 1024:.1f} KB')
    print(f'BCE events: {len(bce_events)}')
    print(f'CE events:  {len(ce_events)}')
    print('Done!')


if __name__ == '__main__':
    main()
